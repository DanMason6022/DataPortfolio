import pandas as pd
import openpyxl
from openpyxl.styles import Font


#reading both files
data=pd.read_excel('od.xlsx', header=[1])
data2=pd.read_excel('os.xlsx', header=[1])
#filling empty axial_lengths with the standard value of '26'
data['Axial_Length'].fillna(26.00,inplace=True)
data2['Axial_Length'].fillna(26.00,inplace=True)

#Checking gender column for mispellings:
unique_gender=data['Gender'].unique()
unique_gender2=data2['Gender'].unique()
print(unique_gender, unique_gender2)
#no need to correct 

#Checking diagnosis column for mispellings 
unique_diagnosis=data['Diagnosis'].unique()
unique_diagnosis2=data2['Diagnosis'].unique()
print(unique_diagnosis, unique_diagnosis2 )

#making the diagnois column all lowercase so its easier to manage 
data['Diagnosis']=data['Diagnosis'].str.lower()
data2['Diagnosis']=data2['Diagnosis'].str.lower()

#checking again
unique_diagnois=data['Diagnosis'].unique()
unique_diagnois2=data2['Diagnosis'].unique()
print(unique_diagnois, unique_diagnois2 )

#making a map to fix mispellings
spellmap = {'heal.':'healthy',
            'glau.' : 'glaucoma',}
data['Diagnosis'] = data['Diagnosis'].replace(spellmap)
data2['Diagnosis'] = data2['Diagnosis'].replace(spellmap)

#checking all unique terms in phakic/pseudophakic
unique_Phakic=data['Phakic/Pseudophakic'].unique()
unique_Phakic2=data2['Phakic/Pseudophakic'].unique()
print(unique_Phakic, unique_Phakic2 )
#no further action needed

#finding the mean and std of each column in each file
means=data.mean()
means2=data2.mean()
stds=data.std()
stds2=data2.std()

#scoring the data
data_score= (data-means)/stds
data2_score = (data2-means2)/stds2

#outlining parameters for outliers
outlier = (data_score.abs() >3)
outlier2=(data2_score.abs()>3)
      
#dropping the columns with outliers     
data_clean= data.drop(data.index[outlier.any(axis=1)])
data_clean2 =data2.drop(data2.index[outlier2.any(axis=1)])

#dropping rows with empty columns for both IOD measurements
data_clean=data_clean.dropna(subset=['Perkins','Pneumatic'], how='all')
data_clean2=data_clean2.dropna(subset=['Perkins','Pneumatic'], how='all')

#dropping rows with empty values in any of the following: dioptre 1 and 2, astigmatism, phakic, pachymetry
data_clean=data_clean.dropna(subset=['dioptre_1','dioptre_2','astigmatism','Phakic/Pseudophakic','Pachymetry',])
data_clean2=data_clean2.dropna(subset=['dioptre_1','dioptre_2','astigmatism','Phakic/Pseudophakic','Pachymetry',])

#dropping any rows not present in both datafiles:
matching_ids = set(data_clean['ID']).intersection(set(data_clean2['ID']))
data_clean=data_clean[data_clean['ID'].isin(matching_ids)]
data_clean2=data_clean2[data_clean2['ID'].isin(matching_ids)]

#saving the editted file
data_clean.to_excel('od_cleaned.xlsx',index=False , header=True)
data_clean2.to_excel('os_cleaned.xlsx',index=False, header =True)


#Formatting files with openpyxl
workbook=openpyxl.load_workbook('os_cleaned.xlsx') 
sheet=workbook.active

#inserting an empty row to be input the subheadings 
sheet.insert_rows(2)

#defining the bold font and border style the same as in the merging code
bold_font=Font(bold=True)
thin_border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'),
                                           right=openpyxl.styles.borders.Side(style='thin'),
                                              top=openpyxl.styles.borders.Side(style='thin'),
                                              bottom=openpyxl.styles.borders.Side(style='thin'))

#applying the border and alligning the contents of the first row
target_cells=['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1']
for target in target_cells:
    cell=sheet[target]
    cell.border=thin_border
    cell.alignment=openpyxl.styles.Alignment(horizontal='center')

#alligning the contents of second row
target_cells=['A2','B2','C2','D2','E2','F2','G2','H2','I2','J2','K2','L2','M2']
for target in target_cells:
    cell=sheet[target]
    cell.alignment=openpyxl.styles.Alignment(horizontal='center')
    
#defining the headers
headers = {
    'ID':('A1', bold_font),
    'Age': ('B1', bold_font),
    'Gender': ('C1', bold_font),
    'Diagnosis': ('D1', bold_font),
    'Refractive_Defect': ('E1', bold_font),
    'Phakic/Pseudophakic': ('H1', bold_font),
    'IOP': ('I1', bold_font),
    'Pachymetry': ('K1', bold_font),
    'Axial_Length': ('L1', bold_font),
    'VF_MD': ('M1', bold_font)
    }
#defining the subheaders
headers2= {
   'ID':('A2'),
   'Age':('B2'),
   'Gender':('C2'),
   'Diagnosis':('D2'),
   'dioptre_1':('E2'),
   'dioptre_2':('F2'),
   'astigmatism':('G2'),
   'Phakic/Pseudophakic':('H2'),
   'Pneumatic':('I2'),
   'Perkins':('J2'),
   'Pachymetry': ('K2'),
   'Axial_Length': ('L2'),
   'VF_MD': ('M2')}
    
#inputting the headers
for header, (cell_address, font_style) in headers.items():
    cell = sheet[cell_address]
    cell.value = header
    cell.font = font_style
    #inputting subheaders
for header, (cell_address) in headers2.items():
    cell = sheet[cell_address]
    cell.value = header

#autowidth as in merging code
for column in sheet.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in sheet[column_letter]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except TypeError:
            pass
    new_width = (max_length+2)*1.1 
    sheet.column_dimensions[column_letter].width = new_width

sheet.merge_cells('E1:G1')
sheet.merge_cells('I1:J1')

workbook.save('os_cleaned.xlsx')

#doing the exact same process for the other file
workbook=openpyxl.load_workbook('od_cleaned.xlsx') 
sheet=workbook.active


bold_font=Font(bold=True)
thin_border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'),
                                           right=openpyxl.styles.borders.Side(style='thin'),
                                              top=openpyxl.styles.borders.Side(style='thin'),
                                              bottom=openpyxl.styles.borders.Side(style='thin'))

sheet.insert_rows(2
                  )
target_cells=['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1']
for target in target_cells:
    cell=sheet[target]
    cell.border=thin_border
    cell.alignment=openpyxl.styles.Alignment(horizontal='center')
    
target_cells=['A2','B2','C2','D2','E2','F2','G2','H2','I2','J2','K2','L2','M2']
for target in target_cells:
    cell=sheet[target]
    cell.alignment=openpyxl.styles.Alignment(horizontal='center')
    
headers = {
    'Age': ('B1', bold_font),
    'Gender': ('C1', bold_font),
    'Diagnosis': ('D1', bold_font),
    'Refractive_Defect': ('E1', bold_font),
    'Phakic/Pseudophakic': ('H1', bold_font),
    'IOP': ('I1', bold_font),
    'Pachymetry': ('K1', bold_font),
    'Axial_Length': ('L1', bold_font),
    'VF_MD': ('M1', bold_font)
    }

headers2={
    'ID':('A2'),
    'Age':('B2'),
    'Gender':('C2'),
    'Diagnosis':('D2'),
    'dioptre_1':('E2'),
    'dioptre_2':('F2'),
    'astigmatism':('G2'),
    'Phakic/Pseudophakic':('H2'),
    'Pneumatic':('I2'),
    'Perkins':('J2'),
    'Pachymetry': ('K2'),
    'Axial_Length': ('L2'),
    'VF_MD': ('M2')}

for header, (cell_address, font_style) in headers.items():
    cell = sheet[cell_address]
    cell.value = header
    cell.font = font_style
    
for header, (cell_address) in headers2.items():
    cell = sheet[cell_address]
    cell.value = header


    
sheet.merge_cells('E1:G1')
sheet.merge_cells('I1:J1')

workbook.save('od_cleaned.xlsx')

